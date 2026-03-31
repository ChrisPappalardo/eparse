# -*- coding: utf-8 -*-

"""
excel parser core module
"""

from io import StringIO
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl.utils.cell import get_column_letter

TableRef = Tuple[int, int, str, str]  # r, c, excel RC, value


# NOTE: df[n] df.at[r,c] and df.iloc[r,c] are not all the same
#       only with .iloc is it safe to assume index and column
#       names will be ingored ; use iloc when working with
#       sub-tables, e.g. the output from df_parse_table


def df_find_tables(
    df: pd.DataFrame,
    loose: bool = False,
) -> List[TableRef]:
    """
    finds table corners in a dataframe
    """

    result = []

    # for each row
    for r in range(df.shape[0]):
        # for each col
        for c in range(df.shape[1]):
            isna_left = True if c == 0 else pd.isna(df.at[r, (c - 1)])
            isna_above = True if r == 0 else pd.isna(df.at[(r - 1), c])
            isna_value = pd.isna(df.at[r, c])

            try:
                isna_right = pd.isna(df.at[r, (c + 1)])
                isna_down = pd.isna(df.at[(r + 1), c])
                isna_corner = pd.isna(df.at[(r + 1), (c + 1)])

                # ensure a 2x2 sparse table
                min_size = any(
                    [
                        not isna_right and not isna_down,
                        not isna_right and not isna_corner,
                        not isna_down and not isna_corner,
                    ]
                )

                if not loose:
                    # ensure a 2x2 dense table
                    min_size = all(
                        [
                            not isna_right,
                            not isna_down,
                            not isna_corner,
                        ]
                    )

            except Exception:
                min_size = False

            if all([isna_left, isna_above, not isna_value, min_size]):
                result.append(
                    (
                        r,
                        c,
                        f"{get_column_letter(c+1)}{r+1}",
                        str(df.at[r, c]),
                    )
                )

    return result


def _get_table_bounds(df: pd.DataFrame, r: int, c: int) -> Tuple[int, int, int, int]:
    """
    calculate complete table zone from top-left corner

    returns (r_start, r_end, c_start, c_end)
    """
    r_start = r
    c_start = c

    # find last column by checking first row
    c_end = c + 1
    for col in range(c + 1, df.shape[1]):
        if pd.isna(df.at[r, col]):
            break
        c_end = col + 1

    # find last row where any cell in column range has data
    r_end = r + 1
    for row in range(r + 1, df.shape[0]):
        row_empty = all(pd.isna(df.at[row, col]) for col in range(c_start, c_end))
        if row_empty:
            break
        r_end = row + 1

    return (r_start, r_end, c_start, c_end)


def _is_inside_bounds(candidate: TableRef, bounds: Tuple[int, int, int, int]) -> bool:
    """
    check if candidate table is inside given bounds
    """
    r, c, _, _ = candidate
    r_start, r_end, c_start, c_end = bounds

    return r_start < r < r_end and c_start < c < c_end


def _filter_nested_tables(
    candidates: List[TableRef], df: pd.DataFrame
) -> List[TableRef]:
    """
    filter out tables that are inside larger tables

    larger tables reserve their zone first, nested tables are excluded
    """
    if len(candidates) <= 1:
        return candidates

    # calculate zones and sizes for all candidates
    zones_data = []
    for candidate in candidates:
        r, c, _, _ = candidate
        bounds = _get_table_bounds(df, r, c)
        r_start, r_end, c_start, c_end = bounds
        size = (r_end - r_start) * (c_end - c_start)
        zones_data.append((candidate, bounds, size))

    # sort by size descending (largest tables first)
    zones_data.sort(key=lambda x: x[2], reverse=True)

    # filter: first tables reserve zone, check if later tables are nested
    filtered = []
    occupied = []

    for candidate, bounds, size in zones_data:
        is_nested = any(_is_inside_bounds(candidate, occ) for occ in occupied)

        if not is_nested:
            filtered.append(candidate)
            occupied.append(bounds)

    # restore original order by position (top-left first)
    filtered.sort(key=lambda x: (x[0], x[1]))

    return filtered


def _is_rowspan(df: pd.DataFrame, r: int, c: int) -> bool:
    """
    detect a rowspan label
    """

    try:
        isna_right = pd.isna(df.at[r, (c + 1)])
        isna_down = pd.isna(df.at[(r + 1), c])

        return isna_down and not isna_right

    except KeyError:
        return False


def _has_empty_corner(df: pd.DataFrame, r: int, c: int) -> bool:
    """
    detect an empty corner
    """

    try:
        isna_above = pd.isna(df.at[(r - 1), c])
        isna_up_corner = pd.isna(df.at[(r - 1), (c + 1)])

        return isna_above and not isna_up_corner

    except KeyError:
        return False


def df_parse_table(
    df: pd.DataFrame,
    r: int,
    c: int,
    na_tolerance_r: int = 1,
    na_tolerance_c: int = 1,
    na_strip: bool = True,
) -> pd.DataFrame:
    """
    extract a table from a dataframe for a given r, c position
    """

    # make reference adjustments
    if _is_rowspan(df, r, c):
        c += 1

    if _has_empty_corner(df, r, c):
        r -= 1

    _r = r + 1

    # get ending row
    na_count = 0
    for row in range(r + 1, df.shape[0]):
        if pd.isna(df.at[row, c]):
            na_count += 1
        else:
            na_count = 0
        if na_count == na_tolerance_r:
            break
        _r += 1
    _c = c + 1

    # get ending col
    na_count = 0
    for col in range(c + 1, df.shape[1]):
        if pd.isna(df.at[r, col]):
            na_count += 1
        else:
            na_count = 0
        if na_count == na_tolerance_c:
            break
        _c += 1

    # strip ending na
    if na_strip and df.iloc[r:_r, c:_c].iloc[-1].isna().all():
        _r -= 1
    if na_strip and df.iloc[r:_r, c:_c].iloc[:, ((_c - c) - 1)].isna().all():
        _c -= 1

    return df.iloc[r:_r, c:_c]


def df_normalize_data(data: Dict) -> Dict:
    """
    normalize table data
    """

    result = {}
    ints = ("row", "column")
    strs = (
        "value",
        "type",
        "c_header",
        "r_header",
        "excel_RC",
        "name",
        "sheet",
        "f_name",
    )

    for k in ints:
        if k in data:
            result[k] = int(data[k])

    for k in strs:
        if k in data:
            result[k] = str(data[k])

    if "timestamp" in data:
        if isinstance(data["timestamp"], pd.Timestamp):
            result["timestamp"] = data["timestamp"].to_pydatetime()

    return result


def df_serialize_table(
    df: pd.DataFrame,
    **other_data,
) -> List[Dict]:
    """
    serialize table into a list of dicts with meta data
    """

    column_header = df.iloc[0]
    row_header = df.iloc[:, 0]

    result = []

    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            _r = df.index[r]  # excel df row
            _c = df.columns[c]  # excel df col
            result.append(
                df_normalize_data(
                    {
                        "row": r,
                        "column": c,
                        "value": df.iloc[r, c],
                        "type": type(df.iloc[r, c]),
                        "c_header": column_header.iloc[c],
                        "r_header": row_header.iloc[r],
                        "excel_RC": f"{get_column_letter(_c+1)}{_r+1}",
                        **other_data,
                    }
                )
            )

    return result


def get_df_from_file(
    io: Any,
    loose: bool = True,
    sheet: Iterable = [],
    table: str = None,
    na_tolerance_r: int = 1,
    na_tolerance_c: int = 1,
    na_strip: bool = True,
    exclude_nested: bool = False,
):
    """
    helper function to yield tables from a file
    """

    f = pd.read_excel(
        io,
        sheet_name=list(sheet) or None,
        header=None,
        index_col=None,
    )

    # convert to dict if single sheet
    if type(f) is not dict:
        f = {s: f for s in sheet}

    for s in f.keys():
        tables = df_find_tables(f[s], loose)

        # apply nested table filter if enabled
        if exclude_nested:
            tables = _filter_nested_tables(tables, f[s])

        for r, c, excel_RC, name in tables:
            if table is not None and table.lower() not in name.lower():
                continue

            yield (
                df_parse_table(
                    f[s],
                    r,
                    c,
                    na_tolerance_r,
                    na_tolerance_c,
                    na_strip,
                ),
                excel_RC,
                name,
                s,
            )


def get_table_digest(
    serialized_table: List[Dict],
    table_name: str,
    filename: Optional[str] = None,
    sheet: Optional[str] = None,
) -> str:
    """
    generate a digest that describes a serialized table
    """

    df = pd.DataFrame.from_records(serialized_table)
    rows = len(df["row"].unique())
    cols = len(df["column"].unique())
    c_headers = df["c_header"].unique()
    r_headers = df["r_header"].unique()
    types = df["type"].unique()

    sheet_str = f" in sheet {sheet}" if sheet else ""
    file_str = f" of Excel file {filename}" if filename else ""
    type_str = f' {", ".join([str(t) for t in types])} type(s)'

    digest = (
        f"{table_name} is a table{sheet_str}{file_str} "
        f'with {cols} column(s) having names like {", ".join(c_headers)} '
        f'and {rows} row(s) having names like {", ".join(r_headers)} '
        f"and contains {rows*cols} cells of{type_str}"
    )

    return digest


def html_to_df(
    html: str,
) -> pd.DataFrame:
    """
    helper function to return pandas dataframe from html
    """

    return pd.read_html(
        StringIO(html),
        header=None,
        index_col=None,
    )


def html_to_serialized_data(
    html: str,
    **other_data,
) -> List[Dict]:
    """
    helper function to return serialized data from html
    """

    return df_serialize_table(html_to_df(html)[0], **other_data)
